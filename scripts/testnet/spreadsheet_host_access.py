#!/usr/bin/env python3
"""Run Synergy Testnet host commands using the operator credential workbook.

The workbook is the source of truth for host, SSH user, custom port, and
credential fields. This helper intentionally never prints credential values.
"""

from __future__ import annotations

import argparse
import os
import shlex
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path("/Users/devpup/Desktop/node machine credentials.xlsx")


@dataclass(frozen=True)
class HostRow:
    row_number: int
    node: str
    ssh_command: str
    ssh_user: str
    public_ip: str
    qrpc_port: str
    ws_port: str
    metrics_port: str
    password: str
    passphrase: str


def load_hosts(workbook: Path) -> dict[str, HostRow]:
    wb = load_workbook(workbook, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    hosts: dict[str, HostRow] = {}
    for row_number, row in enumerate(rows[1:], start=2):
        data = {headers[index]: row[index] for index in range(len(headers))}
        node = str(data.get("Node") or "").strip()
        if not node:
            continue
        host = HostRow(
            row_number=row_number,
            node=node,
            ssh_command=str(data.get("Access Via SSH with") or "").strip(),
            ssh_user=str(data.get("SSH User") or "").strip(),
            public_ip=str(data.get("Public IP") or "").strip(),
            qrpc_port=str(data.get("qRPC") or "").strip(),
            ws_port=str(data.get("WS") or "").strip(),
            metrics_port=str(data.get("Metrics") or "").strip(),
            password=str(data.get("User Password") or ""),
            passphrase=str(data.get("SSH Passphrase") or ""),
        )
        hosts[node.lower()] = host
        hosts[node.replace(" ", "").replace("-", "").lower()] = host
    return hosts


def sanitized_host_line(host: HostRow) -> str:
    return (
        f"spreadsheet_row_used=true row={host.row_number} node={host.node} "
        f"ssh={host.ssh_command!r} user={host.ssh_user!r} "
        f"public_ip={host.public_ip!r} qrpc={host.qrpc_port!r} "
        f"ws={host.ws_port!r} metrics={host.metrics_port!r}"
    )


def run_remote(
    host: HostRow,
    remote_command: str,
    timeout: int,
    password_auth: bool = False,
    extra_env: list[str] | None = None,
) -> int:
    if not host.ssh_command:
        print(f"missing SSH command for {host.node}", file=sys.stderr)
        return 2
    if password_auth:
        print(
            "refusing forced password-auth; workbook SSH command must be used exactly",
            file=sys.stderr,
        )
        return 2
    password = host.password or host.passphrase
    env = os.environ.copy()

    # Preserve the exact workbook SSH command and append only remote environment
    # values sourced from the same row plus the shell command.
    ssh_parts = shlex.split(host.ssh_command)
    env_items = [
        f"SYNERGY_SPREADSHEET_ROW={shlex.quote(str(host.row_number))}",
        f"SYNERGY_NODE={shlex.quote(host.node)}",
        f"SYNERGY_QRPC_PORT={shlex.quote(host.qrpc_port)}",
        f"SYNERGY_WS_PORT={shlex.quote(host.ws_port)}",
        f"SYNERGY_METRICS_PORT={shlex.quote(host.metrics_port)}",
    ]
    for item in extra_env or []:
        if "=" not in item:
            raise ValueError(f"remote env must be NAME=VALUE, got {item!r}")
        name, value = item.split("=", 1)
        if not name.replace("_", "").isalnum():
            raise ValueError(f"invalid remote env name {name!r}")
        env_items.append(f"{name}={shlex.quote(value)}")
    remote_env = " ".join(env_items)
    remote_shell = f"env {remote_env} bash -lc {shlex.quote(remote_command)}"
    print(sanitized_host_line(host), flush=True)
    askpass_path = None
    try:
        if password:
            fd, askpass_path = tempfile.mkstemp(prefix="synergy-ssh-askpass-", text=True)
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                handle.write("#!/bin/sh\n")
                handle.write('printf "%s\\n" "$SYNERGY_SSH_SECRET"\n')
            os.chmod(askpass_path, 0o700)
            env["SYNERGY_SSH_SECRET"] = password
            env["SSH_ASKPASS"] = askpass_path
            env["SSH_ASKPASS_REQUIRE"] = "force"
            env.setdefault("DISPLAY", ":0")
        completed = subprocess.run(
            [*ssh_parts, remote_shell],
            env=env,
            text=True,
            timeout=timeout,
            stdin=subprocess.DEVNULL,
        )
        return completed.returncode
    finally:
        env.pop("SYNERGY_SSH_SECRET", None)
        if askpass_path:
            try:
                os.unlink(askpass_path)
            except FileNotFoundError:
                pass


def scp_command_parts(host: HostRow, password_auth: bool = False) -> list[str]:
    ssh_parts = shlex.split(host.ssh_command)
    if not ssh_parts or ssh_parts[0] != "ssh":
        raise ValueError(f"unsupported SSH command for scp conversion: {host.ssh_command}")
    scp_parts = ["scp"]
    if password_auth:
        if host.ssh_command.strip() == "ssh synergyvps":
            raise ValueError("refusing password-auth SSH options for raw synergyvps alias")
        scp_parts.extend(
            [
                "-o",
                "PreferredAuthentications=password",
                "-o",
                "PubkeyAuthentication=no",
                "-o",
                "IdentitiesOnly=yes",
            ]
        )
    index = 1
    remote_target = None
    while index < len(ssh_parts):
        part = ssh_parts[index]
        if part == "-p" and index + 1 < len(ssh_parts):
            scp_parts.extend(["-P", ssh_parts[index + 1]])
            index += 2
            continue
        if part.startswith("-"):
            scp_parts.append(part)
            index += 1
            continue
        remote_target = part
        index += 1
    if remote_target is None:
        raise ValueError(f"missing remote target in SSH command: {host.ssh_command}")
    return scp_parts, remote_target


def transfer_file(
    host: HostRow,
    direction: str,
    local_path: Path,
    remote_path: str,
    timeout: int,
    password_auth: bool = False,
) -> int:
    if password_auth:
        print(
            "refusing forced password-auth; workbook SSH command must be used exactly",
            file=sys.stderr,
        )
        return 2
    password = host.password or host.passphrase
    env = os.environ.copy()
    ssh_parts = shlex.split(host.ssh_command)
    askpass_path = None
    stdin_handle = None
    stdout_handle = None
    print(sanitized_host_line(host), flush=True)
    try:
        if password:
            fd, askpass_path = tempfile.mkstemp(prefix="synergy-ssh-askpass-", text=True)
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                handle.write("#!/bin/sh\n")
                handle.write('printf "%s\\n" "$SYNERGY_SSH_SECRET"\n')
            os.chmod(askpass_path, 0o700)
            env["SYNERGY_SSH_SECRET"] = password
            env["SSH_ASKPASS"] = askpass_path
            env["SSH_ASKPASS_REQUIRE"] = "force"
            env.setdefault("DISPLAY", ":0")
        if direction == "upload":
            stdin_handle = local_path.open("rb")
            remote_shell = f"cat > {shlex.quote(remote_path)}"
            stdout_target = None
        else:
            local_path.parent.mkdir(parents=True, exist_ok=True)
            stdout_handle = local_path.open("wb")
            remote_shell = f"cat {shlex.quote(remote_path)}"
            stdout_target = stdout_handle
        completed = subprocess.run(
            [*ssh_parts, remote_shell],
            env=env,
            text=False,
            timeout=timeout,
            stdin=stdin_handle or subprocess.DEVNULL,
            stdout=stdout_target,
        )
        return completed.returncode
    finally:
        if stdin_handle:
            stdin_handle.close()
        if stdout_handle:
            stdout_handle.close()
        env.pop("SYNERGY_SSH_SECRET", None)
        if askpass_path:
            try:
                os.unlink(askpass_path)
            except FileNotFoundError:
                pass


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    subparsers = parser.add_subparsers(dest="command", required=True)

    inventory = subparsers.add_parser("inventory")
    inventory.add_argument("--nodes", nargs="*")

    run = subparsers.add_parser("run")
    run.add_argument("node")
    run.add_argument("remote_command")
    run.add_argument("--timeout", type=int, default=60)
    run.add_argument("--remote-env", action="append", default=[])
    run.add_argument(
        "--password-auth",
        action="store_true",
        help="Suppress local SSH keys for non-synergyvps hosts when exact SSH fails from key flood.",
    )

    run_file = subparsers.add_parser("run-file")
    run_file.add_argument("node")
    run_file.add_argument("script_path", type=Path)
    run_file.add_argument("--timeout", type=int, default=60)
    run_file.add_argument("--remote-env", action="append", default=[])
    run_file.add_argument(
        "--password-auth",
        action="store_true",
        help="Suppress local SSH keys for non-synergyvps hosts when exact SSH fails from key flood.",
    )

    download = subparsers.add_parser("download")
    download.add_argument("node")
    download.add_argument("remote_path")
    download.add_argument("local_path", type=Path)
    download.add_argument("--timeout", type=int, default=120)
    download.add_argument("--password-auth", action="store_true")

    upload = subparsers.add_parser("upload")
    upload.add_argument("node")
    upload.add_argument("local_path", type=Path)
    upload.add_argument("remote_path")
    upload.add_argument("--timeout", type=int, default=120)
    upload.add_argument("--password-auth", action="store_true")

    args = parser.parse_args()
    hosts = load_hosts(args.workbook)

    if args.command == "inventory":
        selected = args.nodes or sorted({host.node for host in hosts.values()})
        for node in selected:
            host = hosts.get(node.lower()) or hosts.get(
                node.replace(" ", "").replace("-", "").lower()
            )
            if host is None:
                print(f"missing workbook row for node={node}", file=sys.stderr)
                return 2
            print(sanitized_host_line(host))
        return 0

    if args.command == "run":
        host = hosts.get(args.node.lower()) or hosts.get(
            args.node.replace(" ", "").replace("-", "").lower()
        )
        if host is None:
            print(f"missing workbook row for node={args.node}", file=sys.stderr)
            return 2
        return run_remote(
            host,
            args.remote_command,
            args.timeout,
            args.password_auth,
            args.remote_env,
        )

    if args.command == "run-file":
        host = hosts.get(args.node.lower()) or hosts.get(
            args.node.replace(" ", "").replace("-", "").lower()
        )
        if host is None:
            print(f"missing workbook row for node={args.node}", file=sys.stderr)
            return 2
        return run_remote(
            host,
            args.script_path.read_text(),
            args.timeout,
            args.password_auth,
            args.remote_env,
        )

    if args.command in {"download", "upload"}:
        host = hosts.get(args.node.lower()) or hosts.get(
            args.node.replace(" ", "").replace("-", "").lower()
        )
        if host is None:
            print(f"missing workbook row for node={args.node}", file=sys.stderr)
            return 2
        return transfer_file(
            host,
            args.command,
            args.local_path,
            args.remote_path,
            args.timeout,
            args.password_auth,
        )

    return 2


if __name__ == "__main__":
    raise SystemExit(main())
